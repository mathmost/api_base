# coding: utf-8
import os
import ujson
from pathlib2 import Path
from openpyxl import load_workbook


class OperationExcel:
    """excel操作"""

    def __init__(self, file_path):
        """
        1. 考虑加载表格时，会跨表操作，所以sheet_name不写在初始脚本中
        2. 如果有对表格颜色、表格合并等操作自行添加
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError("文件不存在")
        if Path(file_path).suffix in [".xls", ".xlsx"]:
            self.filepath = file_path
        else:
            raise RuntimeError("文件格式必须为xls或xlsx结尾")
        self.workbook = load_workbook(self.filepath)

    def sheet_names(self):
        """返回所有的sheets"""
        return self.workbook.sheetnames

    def sheet_row_cow(self, sheet_name):
        """获取sheet最大行和最大列"""
        try:
            sheet_val = self.workbook[sheet_name]
        except KeyError:
            raise KeyError("表名不存在")
        row_column = {
            "msg": "success",
            "row": sheet_val.max_row,
            "column": sheet_val.max_column
        }

        return row_column

    def sheet_row_cow_val(self, sheet_name):
        """获取sheet所有数据, 返回二维数组"""
        import numpy as np

        try:
            sheet_val = self.workbook[sheet_name]
        except KeyError:
            raise KeyError("表名不存在")
        # 返回A1, B1, C1这样的顺序
        ls_dic = []
        for row in sheet_val.rows:
            ls_lst = []
            for cell in row:
                if cell.value is not None:
                    ls_lst.append(cell.value)
            ls_dic.append(ls_lst)
        row_column = {
            "msg": "success",
            "all_value": np.array(ls_dic, dtype=object)
        }

        return row_column

    def write_excel(self, sheet_name, wri_row: int, wri_column: int, wri_value):
        """
        写入数据
        :param sheet_name: 表名
        :param wri_row: 写入的行数
        :param wri_column: 写入的列数
        :param wri_value: 写入的值
        """
        sheet_val = self.workbook[sheet_name]
        if type(wri_value) == dict:
            wri_value = ujson.dumps(wri_value, ensure_ascii=False)
        sheet_val.cell(row=wri_row, column=wri_column, value=wri_value)

    def save_data(self):
        """写入之后保存数据"""
        self.workbook.save(self.filepath)
